import psycopg2
from openpyxl import load_workbook
from dateutil.parser import parse
import datetime
from trs_path import expand_paths, abbrev_paths, trs_path_sortkey
import re

from const import *


def update_map(update):

    update_wb = False
    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        # Get maptype abbreviations inddexed by maptype
        cur.execute("""
            SELECT array_agg(ARRAY[t.maptype, t.abbrev]) maptypes FROM {table_maptype} t;
        """.format(table_maptype=TABLE_PROD_MAPTYPE))
        maptypes = dict(cur.fetchone()[0])

        # Get the surveyors and hollins aliases
        surveyors = []
        hollins_surveyor = {}
        wb = load_workbook(XLSX_DATA_SURVEYOR, read_only=True)
        ws = wb.active
        ws_cols = list(c.value for c in ws[1])
        for ws_row in ws.iter_rows(min_row=2):
            ws_value = dict(zip((k.lower() for k in ws_cols), (c.value for c in ws_row)))
            surveyors.append(ws_value['fullname'])
            hollins_surveyor[ws_value['hollins_fullname']] = ws_value['fullname']
        wb.close()

        # Get the parcel map numbers indexed by book/page
        pm_number = {}
        wb = load_workbook(XLSX_DATA_PM, read_only=True)
        ws = wb.active
        ws_cols = list(c.value for c in ws[1])
        for ws_row in ws.iter_rows(min_row=2):
            ws_value = dict(zip((k.lower() for k in ws_cols), (c.value for c in ws_row)))
            book_page = '{1} {0} {2}'.format(ws_value['maptype'], ws_value['book'], ws_value['page'])
            pm_number[book_page] = '(PM%s)' % ws_value['pm_number']
        wb.close()

        # Get the subdivision tract numbers indexed by book/page
        tract_number = {}
        wb = load_workbook(XLSX_DATA_TRACT, read_only=True)
        ws = wb.active
        ws_cols = list(c.value for c in ws[1])
        for ws_row in ws.iter_rows(min_row=2):
            ws_value = dict(zip((k.lower() for k in ws_cols), (c.value for c in ws_row)))
            book_page = '{1} {0} {2}'.format(ws_value['maptype'], ws_value['book'], ws_value['page'])
            tract_number[book_page] = '(TR%s)' % ws_value['tract_number']
        wb.close

        # Get the map update
        wb = load_workbook(XLSX_DATA_MAP)
        ws = wb['update']
        ws_cols = list(c.value for c in ws[1])

        sql = """
            WITH q1 AS (
                SELECT m.id map_id
                FROM {table_map} m
                JOIN {table_maptype} t ON t.id = m.maptype_id
                WHERE t.maptype = %s AND m.book = %s AND m.page = %s
            ), q2 AS (
                SELECT q1.map_id, array_remove(array_agg(p.trs_path::text), NULL) trs_paths
                FROM q1
                LEFT JOIN {table_trs_path} p USING (map_id)
                GROUP BY q1.map_id
            ), q3 AS (
                SELECT q1.map_id, array_remove(array_agg(s.fullname), NULL) surveyors
                FROM q1
                LEFT JOIN {table_signed_by} sb USING (map_id)
                LEFT JOIN {table_surveyor} s ON s.id = sb.surveyor_id
                GROUP BY q1.map_id
            )
            SELECT q1.map_id,
                m.npages, m.recdate, m.client, m.description, m.note,
                q2.trs_paths, q3.surveyors
            FROM {table_map} m
            JOIN q1 ON q1.map_id = m.id
            LEFT JOIN q2 ON q2.map_id = m.id
            LEFT JOIN q3 ON q3.map_id = m.id
            ;
        """.format(
            table_map=TABLE_PROD_MAP,
            table_maptype=TABLE_PROD_MAPTYPE,
            table_trs_path=TABLE_PROD_TRS_PATH,
            table_signed_by=TABLE_PROD_SIGNED_BY,
            table_surveyor=TABLE_PROD_SURVEYOR
        )

        for ws_row in ws.iter_rows(min_row=2):
            ws_value = dict(zip((k.lower() for k in ws_cols), (c.value for c in ws_row)))
            ws_cell = dict(zip((k.lower() for k in ws_cols), ws_row))

            if any(v is None for v in (ws_value['maptype'], ws_value['book'], ws_value['page'])):
                if all(c.value is None for c in ws_row):
                    continue    # Blank lines are OK
                print('ERROR: Missing maptype/book/page: row=%d' % (ws_row[0].row))
                exit(-1)

            # Validate maptype/book/page
            if ws_value['maptype'] not in maptypes:
                print('ERROR: Bad maptype: row=%d, maptype=%s' % (ws_row[0].row, str(ws_value['maptype'])))
                exit(-1)
            ws_maptype = ws_value['maptype']
            if type(ws_value['book']) is int:
                ws_book = ws_value['book']
            else:
                try:
                    ws_cell['book'].value = ws_book = int(ws_value['book'])
                except ValueError as err:
                    print('ERROR: Bad book number: row=%d, book=%s' % (ws_row[0].row, str(ws_value['book'])))
                    exit(-1)
            if type(ws_value['page']) is int:
                ws_page = ws_value['page']
            else:
                try:
                    ws_cell['page'].value = ws_page = int(ws_value['page'])
                except ValueError as err:
                    print('ERROR: Bad page number: row=%d, page=%s' % (ws_row[0].row, str(ws_value['page'])))
                    exit(-1)
            book_page = '%d %s %d' % (ws_book, ws_maptype, ws_page)

            # Validate any trs path spec
            try:
                if ws_value['trs_paths']:
                    ws_paths = expand_paths(re.split(';?\s+', ws_value['trs_paths']))
                else:
                    ws_paths = []
            except ValueError as err:
                print('ERROR: %s: %s' % (book_page, err))
                exit(-1)

            # Validate recdate
            if ws_value['recdate'] is None:
                ws_recdate = None
            elif type(ws_value['recdate']) is datetime.datetime:
                ws_recdate = ws_value['recdate'].date()
            else:
                try:
                    ws_cell['recdate'].value = ws_recdate = parse(ws_value['recdate']).date()
                except Exception as err:
                    print('ERROR: %s: Bad recdate: %s' % (book_page, err))
                    exit(-1)

            # Validate surveyors
            ws_surveyors = sorted(re.split('\s*,\s*', ws_value['surveyors'])) if ws_value['surveyors'] else []
            for i in range(len(ws_surveyors)):
                if ws_surveyors[i] in hollins_surveyor:
                    # replace hollins fullanme with hummaps fullname
                    ws_surveyors[i] = hollins_surveyor[ws_surveyors[i]]
                    ws_cell['surveyors'].value = ', '.join(ws_surveyors)
                elif ws_surveyors[i] not in surveyors:
                    print('ERROR: %s: Missing surveyor: %s' % (book_page, ws_surveyors[i]))
                    exit(-1)

            # Need to add parcel map/tract numbers to client records
            ws_client = ws_value['client']
            if ws_value['maptype'] == 'Parcel Map':
                if book_page in pm_number:
                    ws_client += ' ' + pm_number[book_page]
                else:
                    print('WARNING: %s: No PM number found.' % book_page)
            elif ws_value['maptype'] == 'Record Map':
                if book_page in tract_number:
                    ws_client += ' ' + tract_number[book_page]
                else:
                    print('WARNING: %s: No Tract number found.' % book_page)

            # Fetch the database record
            cur.execute(sql, (ws_value['maptype'], ws_value['book'], ws_value['page']))

            if cur.rowcount == 0:
                print('%s: No database record.' % book_page)
                continue
            if cur.rowcount > 1:
                print('ERROR: %s: Multiple database records.' % book_page)
                exit(-1)
            db_value = dict(zip((d.name for d in cur.description), cur.fetchone()))

            # Compare worksheet and database data
            db_paths = sorted(db_value['trs_paths'], key=trs_path_sortkey)
            if ws_paths != db_paths:
                ws_paths = '; '.join(abbrev_paths(ws_paths)) if ws_paths else None
                db_paths = '; '.join(abbrev_paths(db_paths)) if db_paths else None
                print('%s: Compare trs_paths: xl=%s db=%s' % (book_page, ws_paths, db_paths))

            db_surveyors = sorted(db_value['surveyors'])
            if ws_surveyors != db_surveyors:
                print('%s: Compare surveyors: xl=%s db=%s' % (book_page, ws_surveyors, db_surveyors))

            db_recdate = db_value['recdate']
            if ws_recdate != db_recdate:
                print('%s: Compare recdate: xl=%s db=%s' % (book_page, ws_recdate, db_recdate))

            db_client = db_value['client']
            if ws_client != db_client:
                print('%s: Compare client: xl=%s db=%s' % (book_page, ws_client, db_client))

            for c in ('map_id', 'npages', 'description', 'note'):
                if ws_value[c] != db_value[c]:
                    print('%s: Compare %s: xl=%s db=%s' % (book_page, c, str(ws_value[c]), str(db_value[c])))

    if update_wb:
        wb.save(filename=XLSX_DATA_MAP)
    wb.close()


def update_map_save():


    wb = load_workbook(XLSX_DATA_MAP)
    ws = wb['update']
    ws_cols = list(c.value for c in ws[1])

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        # Map maptypes to maptype ids
        cur.execute("""
            SELECT t.maptype, t.id FROM {table_maptype} t;
        """.format(table_maptype=TABLE_PROD_MAPTYPE))
        MAPTYPE_ID = dict(cur)

        # Map surveyor full names to surveyor ids
        cur.execute("""
            SELECT s.fullname, s.id FROM {table_surveyor} s;
        """.format(table_surveyor=TABLE_PROD_SURVEYOR))
        SURVEYOR_ID = dict(cur)

        # Add or update the TRS source id
        cur.execute("""
            INSERT INTO {table_source} AS s (id, description, quality)
            VALUES (%(source_id)s, %(description)s, %(quality)s)
            ON CONFLICT (id) DO UPDATE
            SET description = %(description)s, quality = %(quality)s
            WHERE s.id = %(source_id)s
            ;
        """.format(table_source=TABLE_PROD_SOURCE), TRS_SOURCE)
        con.commit()

        for ws_row in ws.iter_rows(min_row=2):

            if all(c.value is None for c in ws_row):
                continue
            ws_value = dict(zip((k.lower() for k in ws_cols), (c.value for c in ws_row)))

            # Add a maptype_id item to ws_values
            ws_value['maptype_id'] = MAPTYPE_ID[ws_value['maptype']]

            # Add PM/Tract number to client record
            book_page = '%d %s %d' % (ws_value['book'], ws_value['maptype'], ws_value['page'])
            if ws_value['maptype'] == 'Parcel Map' and book_page in pm_number:
                ws_value['client'] += ' ' + pm_number[book_page]
            elif ws_value['maptype'] == 'Record Map' and  book_page in tract_number:
                ws_value['client'] += ' ' + tract_number[book_page]

            # Add the new map record
            cur.execute("""
                INSERT INTO {table_map} AS m (
                    id, maptype_id, book, page, npages, recdate, client, description, note
                ) VALUES (
                    %(map_id)s, %(maptype_id)s, %(book)s, %(page)s, %(npages)s,
                    %(recdate)s, %(client)s, %(description)s, %(note)s
                )
                ON CONFLICT (id) DO UPDATE
                SET maptype_id = %(maptype_id)s, book = %(book)s, page = %(page)s, npages = %(npages)s,
                    recdate = %(recdate)s, client = %(client)s, description = %(description)s, note = %(note)s
                WHERE m.id = %(map_id)s
                ;
            """.format(table_map=TABLE_PROD_MAP), ws_value)

            # Delete any existing trs path records for the map
            cur.execute("""
                DELETE FROM {table_trs_path} WHERE map_id = %(map_id)s;
            """.format(table_trs_path=TABLE_PROD_TRS_PATH), ws_value)

            # Add trs_path records
            if ws_value['trs_paths']:
                map_id = ws_value['map_id']
                source_id = TRS_SOURCE['source_id']
                paths = expand_paths(re.split(';?\s+', ws_value['trs_paths']))
                cur.executemany("""
                    INSERT INTO {table_trs_path} (map_id, trs_path, source_id) VALUES (%s, %s, %s);
                """.format(table_trs_path=TABLE_PROD_TRS_PATH), ((map_id, path, source_id) for path in paths))

            # Delete any existing signed_by records for the map
            cur.execute("""
                DELETE FROM {table_signed_by} WHERE map_id = %(map_id)s;
            """.format(table_signed_by=TABLE_PROD_SIGNED_BY), ws_value)

            # Add signed_by records
            if ws_value['surveyors']:
                map_id = ws_value['map_id']
                fullnames = sorted(re.split(',\s*', ws_value['surveyors']))
                surveyor_ids = list(SURVEYOR_ID[fullname] for fullname in fullnames)
                cur.executemany("""
                    INSERT INTO {table_signed_by} (map_id, surveyor_id) VALUES (%s, %s);
                """.format(table_signed_by=TABLE_PROD_SIGNED_BY),
                    ((map_id, surveyor_id) for surveyor_id in surveyor_ids)
                )

            con.commit()


if __name__ == '__main__':

    update_map(update=False)
