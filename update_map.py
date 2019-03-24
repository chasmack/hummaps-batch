import psycopg2
from openpyxl import load_workbook
from dateutil.parser import parse
import datetime
from trs_path import expand_paths, abbrev_paths, trs_path_sortkey
import re

from const import *


def update_map(update_db):

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        # Get the update workbook
        wb = load_workbook(XLSX_DATA_MAP, read_only=True)

        # Get maptype ids indexed by maptype
        cur.execute("""
            SELECT t.maptype, t.id FROM {table_maptype} t;
        """.format(table_maptype=TABLE_PROD_MAPTYPE))
        maptype_id = dict(cur)

        # Map surveyor full names to surveyor ids
        cur.execute("""
                   SELECT s.fullname, s.id FROM {table_surveyor} s;
               """.format(table_surveyor=TABLE_PROD_SURVEYOR))
        surveyor_id = dict(cur)

        # Get the parcel map numbers indexed by book-page
        pm_number = {}
        ws = wb['pm_number']
        xl_cols = list(c.value for c in ws[1])
        for xl_row in ws.iter_rows(min_row=2):
            xl_rec = dict(zip((k.lower() for k in xl_cols), (c.value for c in xl_row)))
            book_page = '%s-%s' % (xl_rec['book'], xl_rec['page'])
            pm_number[book_page] = xl_rec['pm_number']

        # Get the subdivision tract numbers indexed by book-page
        tract_number = {}
        ws = wb['tract_number']
        xl_cols = list(c.value for c in ws[1])
        for xl_row in ws.iter_rows(min_row=2):
            xl_rec = dict(zip((k.lower() for k in xl_cols), (c.value for c in xl_row)))
            book_page = '%s-%s' % (xl_rec['book'], xl_rec['page'])
            tract_number[book_page] = xl_rec['tract_number']

        # Query to retreive map info, surveyors and trs paths
        sql = """
            SELECT m.id map_id, m.maptype_id, m.book, m.page, m.npages,
                m.recdate, m.client, m.description, m.note,
                array_remove(array_agg(DISTINCT p.trs_path::text), NULL) trs_paths,
                array_remove(array_agg(DISTINCT sb.surveyor_id), NULL) surveyor_ids
            FROM {table_map} m
            LEFT JOIN {table_signed_by} sb ON sb.map_id = m.id
            LEFT JOIN {table_trs_path} p ON p.map_id = m.id
            WHERE m.id = %(map_id)s
            GROUP BY m.id
            ;
        """.format(
            table_map=TABLE_PROD_MAP,
            table_maptype=TABLE_PROD_MAPTYPE,
            table_trs_path=TABLE_PROD_TRS_PATH,
            table_signed_by=TABLE_PROD_SIGNED_BY,
            table_surveyor=TABLE_PROD_SURVEYOR
        )

        # Lists of database records to update
        map_update = []
        trs_path_delete = []
        trs_path_insert = []
        signed_by_delete = []
        signed_by_insert = []

        # Get the map update
        ws = wb['update']
        xl_cols = list(c.value for c in ws[1])

        for xl_row in ws.iter_rows(min_row=2):

            xl_rec = dict(zip((k.lower() for k in xl_cols), (c.value for c in xl_row)))
            if xl_rec['map_id'] is None:
                continue

            # Add a maptype_id value
            xl_rec['maptype_id'] = maptype_id[xl_rec['maptype']] if xl_rec['maptype'] is not None else None

            # Expand trs path specs into a list
            if xl_rec['trs_paths']:
                xl_rec['trs_paths'] = expand_paths(re.split(';?\s+', xl_rec['trs_paths']))
            else:
                xl_rec['trs_paths'] = []

            # Split surveyors into a list of surveyor ids
            if xl_rec['surveyors']:
                xl_rec['surveyor_ids'] = [surveyor_id[fullname] for fullname in re.split(',\s+', xl_rec['surveyors'])]
            else:
                xl_rec['surveyor_ids'] = []

            # Convert date strings and datetimes to dates
            if type(xl_rec['recdate']) is str:
                xl_rec['recdate'] = parse(xl_rec['recdate']).date()
            elif type(xl_rec['recdate']) is datetime.datetime:
                xl_rec['recdate'] = xl_rec['recdate'].date()

            # Add parcel map/tract numbers to client records
            if all(xl_rec[col] for col in ('maptype', 'book', 'page')):
                book_page = '%s-%s' % (xl_rec['book'], xl_rec['page'])
                if xl_rec['maptype'] == 'Parcel Map' and book_page in pm_number:
                    xl_rec['client'] += ' (PM%s)' % pm_number[book_page]
                elif xl_rec['maptype'] == 'Record Map' and book_page in tract_number:
                    xl_rec['client'] +=  ' (TR%s)' % tract_number[book_page]

            # Fetch the database record
            cur.execute(sql, xl_rec)

            map_id = xl_rec['map_id']
            source_id = TRS_SOURCE['source_id']

            if cur.rowcount == 0:
                # Add new records
                map_update.append(xl_rec)
                trs_path_insert += ((map_id, path, source_id) for path in xl_rec['trs_paths'])
                signed_by_insert += ((map_id, surveyor_id) for surveyor_id in xl_rec['surveyor_ids'])

            else:
                # Compare excel and database records
                db_rec = dict(zip((d.name for d in cur.description), cur.fetchone()))

                cols = ('maptype_id', 'book', 'page', 'npages', 'recdate', 'client', 'description', 'note')
                if any(xl_rec[col] != db_rec[col] for col in cols):
                    map_update.append(xl_rec)

                xl_paths = set(xl_rec['trs_paths'])
                db_paths = set(db_rec['trs_paths'])
                trs_path_delete += ((map_id, path) for path in db_paths - xl_paths)
                trs_path_insert += ((map_id, path, source_id) for path in xl_paths - db_paths)

                xl_surveyor_ids = set(xl_rec['surveyor_ids'])
                db_surveyor_ids = set(db_rec['surveyor_ids'])
                signed_by_delete += ((map_id, surveyor_id) for surveyor_id in db_surveyor_ids - xl_surveyor_ids)
                signed_by_insert += ((map_id, surveyor_id) for surveyor_id in xl_surveyor_ids - db_surveyor_ids)

        wb.close()

        if update_db:
            # Add or update the TRS source id
            cur.execute("""
                INSERT INTO {table_source} AS s (id, description, quality)
                VALUES (%(source_id)s, %(description)s, %(quality)s)
                ON CONFLICT (id) DO UPDATE
                SET description = %(description)s, quality = %(quality)s
                WHERE s.id = %(source_id)s
                ;
            """.format(table_source=TABLE_PROD_SOURCE), TRS_SOURCE)
            con.commit()

        if map_update and update_db:
            # Insert/update the map records
            cur.executemany("""
                INSERT INTO {table_map} AS m (
                    id, maptype_id, book, page, npages, recdate, client, description, note
                ) VALUES (
                    %(map_id)s, %(maptype_id)s, %(book)s, %(page)s, %(npages)s,
                    %(recdate)s, %(client)s, %(description)s, %(note)s
                )
                ON CONFLICT (id) DO UPDATE
                SET maptype_id = %(maptype_id)s, book = %(book)s, page = %(page)s, npages = %(npages)s,
                    recdate = %(recdate)s, client = %(client)s, description = %(description)s, note = %(note)s
                WHERE m.id = %(map_id)s
                ;
            """.format(table_map=TABLE_PROD_MAP), map_update)
            con.commit()

            print('INSERT/UPDATE map: %d row%s' % (cur.rowcount, '' if cur.rowcount == 1 else 's'))

        elif map_update:
            print('INSERT/UPDATE map:')
            for i in range(len(map_update)):
                print('[%d] %s' % (i + 1, map_update[i]))

        if trs_path_delete and update_db:
            # Delete trs path records
            cur.executemany("""
                DELETE FROM {table_trs_path}
                WHERE map_id = %s AND trs_path = %s;
            """.format(table_trs_path=TABLE_PROD_TRS_PATH), trs_path_delete)
            con.commit()

            print('DELETE trs_path: %d row%s' % (cur.rowcount, '' if cur.rowcount == 1 else 's'))

        elif trs_path_delete:
            print('DELETE trs_path:')
            for i in range(len(trs_path_delete)):
                print('[%d] %s' % (i + 1, trs_path_delete[i]))

        if trs_path_insert and update_db:
            # Add trs_path records
            cur.executemany("""
                INSERT INTO {table_trs_path} (map_id, trs_path, source_id) VALUES (%s, %s, %s);
            """.format(table_trs_path=TABLE_PROD_TRS_PATH), trs_path_insert)
            con.commit()

            print('INSERT trs_path: %d row%s' % (cur.rowcount, '' if cur.rowcount == 1 else 's'))

        elif trs_path_insert:
            print('INSERT trs_path:')
            for i in range(len(trs_path_insert)):
                print('[%d] %s' % (i + 1, trs_path_insert[i]))

        if signed_by_delete and update_db:
            # Delete signed_by records
            cur.executemany("""
                DELETE FROM {table_signed_by}
                WHERE map_id = %s AND surveyor_id = %s;
            """.format(table_signed_by=TABLE_PROD_SIGNED_BY), signed_by_delete)
            con.commit()

            print('DELETE signed_by: %d row%s' % (cur.rowcount, '' if cur.rowcount == 1 else 's'))

        elif signed_by_delete:
            print('DELETE signed_by:')
            for i in range(len(signed_by_delete)):
                print('[%d] %s' % (i + 1, signed_by_delete[i]))

        if signed_by_insert and update_db:
            # Add signed_by records
            cur.executemany("""
                INSERT INTO {table_signed_by} (map_id, surveyor_id) VALUES (%s, %s);
            """.format(table_signed_by=TABLE_PROD_SIGNED_BY), signed_by_insert)
            con.commit()

            print('INSERT signed_by: %d row%s' % (cur.rowcount, '' if cur.rowcount == 1 else 's'))

        elif signed_by_insert:
            print('INSERT signed_by:')
            for i in range(len(signed_by_insert)):
                print('[%d] %s' % (i + 1, signed_by_insert[i]))

        updates = (map_update, trs_path_delete, trs_path_insert, signed_by_delete, signed_by_insert)
        if all(len(l) == 0 for l in updates):
            print('Nothing to do.')


if __name__ == '__main__':

    update_map(update_db=False)
