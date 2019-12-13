import psycopg2
from openpyxl import load_workbook
import time
import re

from trs_path import expand_paths

from const import *


def update_map():

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        wb = load_workbook(XLSX_DATA_UPDATE, read_only=True)

        # Insert new map records from the XLSX update
        print('INSERT INTO {table_prod_map} ...'.format(table_prod_map=TABLE_PROD_MAP))

        ws = wb[XLSX_SHEET_UPDATE_NEW]
        column_names = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        rows = [dict(zip(column_names, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
        cur.executemany("""
            INSERT INTO {table_prod_map} (
                id, maptype_id, book, page, npages, recdate, client, description, note
            )
            SELECT
                %(MAP_ID)s, t.id, %(BOOK)s, %(PAGE)s, %(NPAGES)s,
                CAST(%(RECDATE)s AS DATE), %(CLIENT)s, %(DESCRIPTION)s, %(NOTE)s
            FROM {table_prod_maptype} t
            WHERE t.maptype = %(MAPTYPE)s
            ;
        """.format(
            table_prod_map=TABLE_PROD_MAP,
            table_prod_maptype=TABLE_PROD_MAPTYPE
        ), rows)
        con.commit()
        print('%d rows effected.' % cur.rowcount)

        # Update existing map records from the XLSX update
        print('UPDATE {table_prod_map} ...'.format(table_prod_map=TABLE_PROD_MAP))

        ws = wb[XLSX_SHEET_UPDATE_MODIFIED]
        column_names = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        rows = [dict(zip(column_names, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
        cur.executemany("""
            UPDATE {table_prod_map} m SET (
                maptype_id, book, page, npages, recdate, client, description, note
            ) = (
                SELECT
                    t.id, %(BOOK)s, %(PAGE)s, %(NPAGES)s,
                    CAST(%(RECDATE)s AS DATE), %(CLIENT)s, %(DESCRIPTION)s, %(NOTE)s
                FROM {table_prod_maptype} t
                WHERE t.maptype = %(MAPTYPE)s
            )
            WHERE m.id = %(MAP_ID)s
            ;
        """.format(
            table_prod_map=TABLE_PROD_MAP,
            table_prod_maptype=TABLE_PROD_MAPTYPE
        ), rows)
        con.commit()
        print('%d rows effected.' % cur.rowcount)

        wb.close()


def update_trs():

        with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

            wb = load_workbook(XLSX_DATA_UPDATE, read_only=True)

            # Add the trs source id
            cur.execute("""
                INSERT INTO {table_prod_source} (id, description, quality)
                VALUES (%(SOURCE_ID)s, %(DESCRIPTION)s, %(QUALITY)s)
                ON CONFLICT (id) DO UPDATE SET
                  description = %(DESCRIPTION)s,
                  quality = %(QUALITY)s
                ;
            """.format(table_prod_source=TABLE_PROD_SOURCE), TRS_SOURCE)
            con.commit()

            # Insert new trs records from the XLSX update
            print('INSERT INTO {table_prod_trs_path} ...'.format(table_prod_trs_path=TABLE_PROD_TRS_PATH))

            ws = wb[XLSX_SHEET_UPDATE_NEW]
            column_names = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

            rowcount = 0
            for row in [dict(zip(column_names, row)) for row in ws.iter_rows(min_row=2, values_only=True)]:
                for path in expand_paths(re.split('\s*;\s*', row['TRS_PATHS'])):
                    cur.execute("""
                        INSERT INTO {table_prod_trs_path} (map_id, trs_path, source_id) VALUES (%s, %s, %s);
                    """.format(
                        table_prod_trs_path=TABLE_PROD_TRS_PATH
                    ), [row['MAP_ID'], path, TRS_SOURCE['SOURCE_ID']])
                    rowcount += 1
            con.commit()
            print('%d rows effected.' % rowcount)

            # Update existing trs records from the XLSX update
            ws = wb[XLSX_SHEET_UPDATE_MODIFIED]
            column_names = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            rows = [dict(zip(column_names, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

            print('DELETE FROM {table_prod_trs_path} ...'.format(table_prod_trs_path=TABLE_PROD_TRS_PATH))
            cur.executemany("""
                DELETE FROM {table_prod_trs_path} WHERE map_id = %(MAP_ID)s;
            """.format(table_prod_trs_path=TABLE_PROD_TRS_PATH), rows)
            con.commit()
            print('%d rows effected.' % cur.rowcount)

            rowcount = 0
            print('INSERT INTO {table_prod_trs_path} ...'.format(table_prod_trs_path=TABLE_PROD_TRS_PATH))
            for row in [dict(zip(column_names, row)) for row in ws.iter_rows(min_row=2, values_only=True)]:
                for path in expand_paths(re.split('\s*;\s*', row['TRS_PATHS'])):
                    cur.execute("""
                        INSERT INTO {table_prod_trs_path} (map_id, trs_path, source_id) VALUES (%s, %s, %s);
                    """.format(
                        table_prod_trs_path=TABLE_PROD_TRS_PATH
                    ), [row['MAP_ID'], path, TRS_SOURCE['SOURCE_ID']])
                    rowcount += 1
            con.commit()
            print('%d rows effected.' % rowcount)

            wb.close()


def update_surveyor():

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        # Insert new surveyor records from the update table
        print('INSERT INTO {table_prod_surveyor} ...'.format(table_prod_surveyor=TABLE_PROD_SURVEYOR))

        cur.execute("""
            INSERT INTO {table_prod_surveyor}
              (fullname, firstname, secondname, thirdname, lastname, suffix, pls, rce)
            SELECT u.fullname, u.firstname, u.secondname, u.thirdname, u.lastname, u.suffix, u.pls, u.rce
            FROM {table_update_surveyor} u
            LEFT JOIN {table_prod_surveyor} s ON s.fullname = u.fullname
            WHERE s.id IS NULL
            ;
        """.format(
            table_prod_surveyor=TABLE_PROD_SURVEYOR,
            table_update_surveyor=TABLE_UPDATE_SURVEYOR
        ))
        con.commit()
        print('%d rows effected.' % cur.rowcount)

        # Update existing surveyor records from the update table
        print('UPDATE {table_prod_surveyor} ...'.format(table_prod_surveyor=TABLE_PROD_SURVEYOR))

        cur.execute("""
            UPDATE {table_prod_surveyor} s SET (
                fullname, firstname, secondname, thirdname, lastname, suffix, pls, rce
            ) = (
                SELECT DISTINCT
                    u.fullname, u.firstname, u.secondname, u.thirdname, u.lastname, u.suffix, u.pls, u.rce
                FROM {table_update_surveyor} u
                WHERE u.fullname = s.fullname
            );
        """.format(
            table_prod_surveyor=TABLE_PROD_SURVEYOR,
            table_update_surveyor=TABLE_UPDATE_SURVEYOR
        ))
        con.commit()
        print('%d rows effected.' % cur.rowcount)

        # Replace all signed_by records for maps in the update.
        print('DELETE {table_prod_signed_by} ...'.format(table_prod_signed_by=TABLE_PROD_SIGNED_BY))

        cur.execute("""
                    DELETE FROM {table_prod_signed_by}
                    WHERE map_id IN (SELECT u.id FROM {table_update_map} u);
                """.format(
            table_prod_signed_by=TABLE_PROD_SIGNED_BY,
            table_update_map=TABLE_UPDATE_MAP
        ))
        con.commit()
        print('%d rows effected.' % cur.rowcount)

        print('INSERT INTO {table_prod_signed_by} ...'.format(table_prod_signed_by=TABLE_PROD_SIGNED_BY))

        # Link "&" separated list of Hollins surveyors in update map records to surveyor records
        # in production using hollins_fullname in the update surveyor table.
        cur.execute("""
            INSERT INTO {table_prod_signed_by} (map_id, surveyor_id)
            SELECT upd_m.id, s.id
            FROM (
              SELECT u.id, regexp_split_to_table(u.surveyors, '\s*&\s*') hollins_fullname
              FROM {table_update_map} u
            ) upd_m
            JOIN {table_update_surveyor} upd_s ON upd_s.hollins_fullname = upd_m.hollins_fullname
            JOIN {table_prod_surveyor} s ON s.fullname = upd_s.fullname
            ;
        """.format(
            table_prod_signed_by=TABLE_PROD_SIGNED_BY,
            table_prod_surveyor=TABLE_PROD_SURVEYOR,
            table_update_map=TABLE_UPDATE_MAP,
            table_update_surveyor=TABLE_UPDATE_SURVEYOR
        ))
        con.commit()
        print('%d rows effected.' % cur.rowcount)


def update_cc():

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        # Insert new cc records from the update table
        print('INSERT INTO {table_prod_cc} ...'.format(table_prod_cc=TABLE_PROD_CC))

        cur.execute("""
            WITH q1 AS (
                SELECT m.id map_id, t.maptype, m.book, m.page, cc.doc_number
                FROM {table_prod_cc} cc
                JOIN {table_prod_map} m ON m.id = cc.map_id
                JOIN {table_prod_maptype} t ON t.id = m.maptype_id
            )
            INSERT INTO {table_prod_cc} (map_id, doc_number, npages)
            SELECT upd_m.id map_id, upd_cc.doc_number, upd_cc.npages
            FROM {table_update_cc} upd_cc
            JOIN {table_update_map} upd_m ON
                upd_m.maptype = upd_cc.maptype AND
                upd_m.book = upd_cc.book AND
                upd_m.page = upd_cc.page
            LEFT JOIN q1 ON
                q1.maptype = upd_cc.maptype AND
                q1.book = upd_cc.book AND
                q1.page = upd_cc.page AND
                q1.doc_number = upd_cc.doc_number
            WHERE q1.map_id IS NULL
            ;
        """.format(
            table_prod_cc=TABLE_PROD_CC,
            table_prod_map=TABLE_PROD_MAP,
            table_prod_maptype=TABLE_PROD_MAPTYPE,
            table_update_cc=TABLE_UPDATE_CC,
            table_update_map=TABLE_UPDATE_MAP
        ))
        con.commit()
        print('%d rows effected.' % cur.rowcount)


if __name__ == '__main__':


    startTime = time.time()

    update_map()
    update_trs()
    update_surveyor()
    update_cc()

    endTime = time.time()
    print('\n{0:.3f} sec'.format(endTime - startTime))
