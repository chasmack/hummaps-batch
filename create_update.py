import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import psycopg2
import re
import time

from const import *
from trs_path import abbrev_paths


def load_update_tables():

    # Load the update tables from Hollins XML data
    print('\nLoad update tables ... ')

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        # Create the update schema
        print('CREATE SCHEMA: {schema_update} ...'.format(schema_update=SCHEMA_UPDATE))
        cur.execute("""
            DROP SCHEMA IF EXISTS {schema_update} CASCADE;
            CREATE SCHEMA {schema_update};
        """.format(schema_update=SCHEMA_UPDATE))

        # Load map data from XML update
        print('CREATE TABLE: {table_update_map} ...'.format(table_update_map=TABLE_UPDATE_MAP))
        cur.execute("""
            CREATE TABLE {table_update_map} (
              id integer PRIMARY KEY,
              maptype text,
              book integer,
              page integer,
              lastpage integer,
              recdate date,
              surveyors text,
              client text,
              description text,
              image text,
              note text
            );
        """.format(table_update_map=TABLE_UPDATE_MAP))

        tree = ET.parse(XML_DATA_UPDATE_MAP)
        root = tree.getroot()
        rows = []
        keys = ('ID', 'maptype', 'BOOK', 'FIRSTPAGE', 'LASTPAGE', 'RECDATE',
                'SURVEYOR', 'DONEFOR', 'DESCRIP', 'Picture', 'Comment')
        for map_rec in root:
            rec = []
            for key in keys:
                elem = map_rec.find(key)
                val = None if elem is None else elem.text
                if key == 'SURVEYOR' and '&' in val:
                    # Multiple surveyors must be ordered by last name, first initial
                    # Fix map_id=14143 with a missing space in the surveyor "A.BONES"
                    surveyors = (re.sub('\.\s*', '. ', s) for s in re.split('\s*&\s*', val))
                    surveyors = sorted(surveyors, key=lambda v: '%s %s' % (v[3:], v[0:2]))
                    val = ' & '.join(surveyors)
                rec.append(val)
            rows.append(rec)
        cur.executemany("""
            INSERT INTO {table_update_map} (
              id, maptype, book, page, lastpage, recdate,
              surveyors, client, description, image, note
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
        """.format(table_update_map=TABLE_UPDATE_MAP), rows)
        con.commit()
        print('INSERT: ' + str(cur.rowcount) + ' rows effected.')

        # Load trs data from XML update
        print('CREATE TABLE: {table_update_trs_path} ...'.format(table_update_trs_path=TABLE_UPDATE_TRS_PATH))
        cur.execute("""
            CREATE TABLE {table_update_trs_path} (
              id serial PRIMARY KEY,
              map_id integer REFERENCES {table_update_map},
              trs_path ltree
            );
        """.format(
            table_update_trs_path=TABLE_UPDATE_TRS_PATH,
            table_update_map=TABLE_UPDATE_MAP
        ))

        rows = []

        # Get full section data from trs
        tree = ET.parse(XML_DATA_UPDATE_TRS)
        root = tree.getroot()
        for trs in root:
            map_id = trs.find('ID').text

            tshp = trs.find('TOWNSHIP').text
            rng = trs.find('RANGE').text
            if tshp == '0' and rng == '0':
                # Special case for unknown location
                continue
            tr = '%s.%s' % (tshp.lstrip('0'), rng)

            secs = trs.find('SECTION').text
            if re.fullmatch(',([1-3]\d,|[1-9],)+', secs) is None:
                print('ERROR (%s): Bad trs SECTION format: %s' % (XML_DATA_UPDATE_TRS, secs))
                exit(1)
            secs = secs.strip(',').split(',')

            for sec in secs:
                rows.append((map_id, '%s.%s' % (tr, sec)))

        # Get subsection data from map
        tree = ET.parse(XML_DATA_UPDATE_MAP)
        root = tree.getroot()

        # Map hollins subsection codes to trs_path subsection codes
        ss_code = dict(zip((str(n) for n in range(1, 37)), 'DCBAEFGHLKJIMNOP'))

        for map_rec in root.findall('map'):
            for c in map_rec:
                if c.tag == 'ID':
                    map_id = c.text
                elif c.tag.startswith('_x0030_'):
                    tshp = c.tag[7:9]
                    rng = c.tag[9:11]
                    sec = c.tag[11:].lstrip('0')
                    subsecs = c.text
                    if subsecs[0] != ',' or subsecs[-1] != ',':
                        print('ERROR (%s): Bad map subsection format (%s): %s' % (
                        XML_DATA_UPDATE_MAP, c.tag[7:], subsecs))
                        exit(1)
                    trs = '.'.join((tshp, rng, sec))
                    for ss in subsecs.strip(',').split(','):
                        rows.append((map_id, '%s.%s' % (trs, ss_code[ss])))

        cur.executemany("""
            INSERT INTO {table_update_trs_path} (map_id, trs_path)
            VALUES (%s, %s);
        """.format(table_update_trs_path=TABLE_UPDATE_TRS_PATH), rows)
        con.commit()
        print('INSERT: ' + str(cur.rowcount) + ' rows effected.')

        # Remove redundant ancestors from trs paths
        cur.execute("""
            WITH q1 AS (
              SELECT map_id, array_agg(trs_path) paths
              FROM {table_update_trs_path}
              WHERE nlevel(trs_path) = 4
              GROUP BY map_id
            )
            DELETE FROM {table_update_trs_path}
            WHERE trs_path.id IN (
              SELECT id
              FROM {table_update_trs_path}
              JOIN q1 USING (map_id)
              WHERE nlevel(trs_path) = 3 AND trs_path @> paths
            );
        """.format(table_update_trs_path=TABLE_UPDATE_TRS_PATH))
        con.commit()
        print('DELETE: ' + str(cur.rowcount) + ' rows effected.')

    print('Done.')


def load_xlsx_tables():

    # Load update tables from XLSX update data
    print('\nLoad XLSX tables and check update ... ')

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        wb = load_workbook(XLSX_DATA_UPDATE, read_only=True)

        print('CREATE TABLE: {table_update_surveyor} ...'.format(table_update_surveyor=TABLE_UPDATE_SURVEYOR))
        ws = wb[XLSX_SHEET_SURVEYOR]
        cur.execute("""
            DROP TABLE IF EXISTS {table_update_surveyor};
            CREATE TABLE {table_update_surveyor} (
              id serial PRIMARY KEY,
              hollins_fullname text,
              fullname text,
              firstname text,
              secondname text,
              thirdname text,
              lastname text,
              suffix text,
              pls text,
              rce text
            );
        """.format(table_update_surveyor=TABLE_UPDATE_SURVEYOR))
        cur.executemany("""
            INSERT INTO {table_update_surveyor}
              (hollins_fullname, fullname, firstname, secondname, thirdname, lastname, suffix, pls, rce)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s);
        """.format(table_update_surveyor=TABLE_UPDATE_SURVEYOR), ws.iter_rows(min_row=2, values_only=True))
        con.commit()
        print('INSERT: ' + str(cur.rowcount) + ' rows effected.')

        print('CREATE TABLE: {table_update_pm_number} ...'.format(table_update_pm_number=TABLE_UPDATE_PM_NUMBER))
        ws = wb[XLSX_SHEET_PM_NUMBER]
        cur.execute("""
            DROP TABLE IF EXISTS {table_update_pm_number};
            CREATE TABLE {table_update_pm_number} (
              id serial PRIMARY KEY,
              maptype text,
              book integer,
              page integer,
              pm_number text,
              note text
            );
        """.format(table_update_pm_number=TABLE_UPDATE_PM_NUMBER))
        cur.executemany("""
            INSERT INTO {table_update_pm_number}
              (maptype, book, page, pm_number, note)
            VALUES (%s, %s, %s, %s, %s);
        """.format(table_update_pm_number=TABLE_UPDATE_PM_NUMBER), ws.iter_rows(min_row=2, values_only=True))
        con.commit()
        print('INSERT: ' + str(cur.rowcount) + ' rows effected.')

        print('CREATE TABLE: {table_update_tract_number} ...'.format(
            table_update_tract_number=TABLE_UPDATE_TRACT_NUMBER))
        cur.execute("""
            DROP TABLE IF EXISTS {table_update_tract_number};
            CREATE TABLE {table_update_tract_number} (
              id serial PRIMARY KEY,
              maptype text,
              book integer,
              page integer,
              tract_number text,
              note text
            );
        """.format(table_update_tract_number=TABLE_UPDATE_TRACT_NUMBER))
        ws = wb[XLSX_SHEET_TRACT_NUMBER]
        cur.executemany("""
            INSERT INTO {table_update_tract_number}
              (maptype, book, page, tract_number, note)
            VALUES (%s, %s, %s, %s, %s);
        """.format(table_update_tract_number=TABLE_UPDATE_TRACT_NUMBER), ws.iter_rows(min_row=2, values_only=True))
        con.commit()
        print('INSERT: ' + str(cur.rowcount) + ' rows effected.')

        print('CREATE TABLE: {table_update_cc} ...'.format(
            table_update_cc=TABLE_UPDATE_CC))
        cur.execute("""
            DROP TABLE IF EXISTS {table_update_cc};
            CREATE TABLE {table_update_cc} (
              id serial PRIMARY KEY,
              maptype text,
              book integer,
              page integer,
              doc_number text,
              npages integer
            );
        """.format(table_update_cc=TABLE_UPDATE_CC))
        ws = wb[XLSX_SHEET_CC]
        cur.executemany("""
            INSERT INTO {table_update_cc}
              (maptype, book, page, doc_number, npages)
            VALUES (%s, %s, %s, %s, %s);
        """.format(table_update_cc=TABLE_UPDATE_CC), ws.iter_rows(min_row=2, values_only=True))
        con.commit()
        print('INSERT: ' + str(cur.rowcount) + ' rows effected.')

        wb.close()

        print('Check for missing surveyors ...')
        cur.execute("""
            WITH q1 AS (
              SELECT m.id map_id, regexp_split_to_table(upper(m.surveyors), ' & ') hollins_fullname
              FROM {table_update_map} m
            )
            SELECT q1.map_id, q1.hollins_fullname
            FROM q1
            LEFT JOIN {table_update_surveyor} s USING (hollins_fullname)
            WHERE s.id IS NULL
            AND q1.hollins_fullname NOT IN ('UNKNOWN', 'U. NKNOWN', 'A. RAINES?')
            ;
        """.format(
            table_update_map=TABLE_UPDATE_MAP,
            table_update_surveyor=TABLE_UPDATE_SURVEYOR
        ))
        for row in cur:
            print('ERROR: Missing surveyor for map id={0}: "{1}"'.format(*row))
        if cur.rowcount > 0:
            exit(1)

        print('Check for missing PM numbers ...')
        cur.execute("""
            SELECT u.id map_id, u.book, u.page
            FROM {table_update_map} u
            LEFT JOIN {table_prod_map} m ON m.id = u.id
            LEFT JOIN {table_update_pm_number} n ON n.book = u.book AND n.page = u.page
            WHERE u.maptype = 'Parcel Map'
            AND m.id IS NULL
            AND n.pm_number is NULL
            ;
        """.format(
            table_update_map=TABLE_UPDATE_MAP,
            table_prod_map=TABLE_PROD_MAP,
            table_update_pm_number=TABLE_UPDATE_PM_NUMBER
        ))
        for row in cur:
            print('WARNING: Missing PM number for map id={0}: {1}PM{2}'.format(*row))

        print('Check for missing TRACT numbers ...')
        cur.execute("""
            SELECT u.id map_id, u.book, u.page
            FROM {table_update_map} u
            LEFT JOIN {table_prod_map} m ON m.id = u.id
            LEFT JOIN {table_update_tract_number} n ON n.book = u.book AND n.page = u.page
            WHERE u.maptype = 'Record Map'
            AND m.id IS NULL
            AND n.tract_number is NULL
            ;
        """.format(
            table_update_map=TABLE_UPDATE_MAP,
            table_prod_map=TABLE_PROD_MAP,
            table_update_tract_number=TABLE_UPDATE_TRACT_NUMBER
        ))
        for row in cur:
            print('WARNING: Missing TRACT number for map id={0}: {1}RM{2}'.format(*row))

        print('Check for new CCs ...')
        cur.execute("""
            WITH q1 AS (
                SELECT m.id map_id, t.maptype, m.book, m.page, cc.doc_number
                FROM {table_prod_cc} cc
                JOIN {table_prod_map} m ON m.id = cc.map_id
                JOIN {table_prod_maptype} t ON t.id = m.maptype_id
            )
            SELECT upd_m.id map_id, upd_m.maptype, upd_m.book, upd_m.page, upd_cc.doc_number, upd_cc.npages
            FROM {table_update_cc} upd_cc
            JOIN {table_update_map} upd_m ON
                upd_m.maptype = upd_cc.maptype AND
                upd_m.book = upd_cc.book AND
                upd_m.page = upd_cc.page
            LEFT JOIN q1 ON
                q1.maptype = upd_cc.maptype AND
                q1.book = upd_cc.book AND
                q1.page = upd_cc.page AND
                q1.doc_number = upd_cc.doc_number
            WHERE q1.map_id IS NULL
            ;
        """.format(
            table_prod_cc=TABLE_PROD_CC,
            table_prod_map=TABLE_PROD_MAP,
            table_prod_maptype=TABLE_PROD_MAPTYPE,
            table_update_cc=TABLE_UPDATE_CC,
            table_update_map=TABLE_UPDATE_MAP
        ))
        for row in cur:
            print('INFO: New CC for {2} {1}s {3}: map_id={0} doc_number={4} npages={5}'.format(*row))

    print('Done.')


def create_update():

    print('\nCreate XLSX udate.')

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        # Add map data to the XLSX update
        wb = load_workbook(XLSX_DATA_UPDATE)

        print('Add NEW map records to XLSX sheet "{0}" ...'.format(XLSX_SHEET_UPDATE_NEW))

        ws = wb[XLSX_SHEET_UPDATE_NEW]
        ws.delete_rows(2, ws.max_row - 1)
        column_names = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        cur.execute("""
            WITH q1 AS (
                SELECT map_id, array_agg(trs_path) paths
                FROM {table_update_trs_path}
                GROUP BY map_id
            )
            SELECT u.id map_id,
                u.maptype, u.book, u.page, u.lastpage - u.page + 1 npages,
                coalesce(q1.paths::text[], ARRAY[]::text[]) trs_paths,
                u.recdate, u.surveyors, u.client, u.description, u.note,
                pm_number, tract_number
            FROM {table_update_map} u
            LEFT JOIN q1 ON q1.map_id = u.id
            LEFT JOIN {table_prod_map} m ON u.id = m.id
            LEFT JOIN {table_update_pm_number} pm
                ON pm.maptype = u.maptype AND pm.book = u.book AND pm.page = u.page
            LEFT JOIN {table_update_tract_number} tr
                ON tr.maptype = u.maptype AND tr.book = u.book AND tr.page = u.page
            WHERE m.id IS NULL
            ORDER BY u.id
            ;
        """.format(
            table_update_map=TABLE_UPDATE_MAP,
            table_update_trs_path=TABLE_UPDATE_TRS_PATH,
            table_update_pm_number=TABLE_UPDATE_PM_NUMBER,
            table_update_tract_number=TABLE_UPDATE_TRACT_NUMBER,
            table_prod_map=TABLE_PROD_MAP
        ))

        for row in cur:
            pm_number, tract_number = row[-2:]
            row = dict(zip(column_names, row))
            row['TRS_PATHS'] = '; '.join(abbrev_paths(row['TRS_PATHS']))
            if row['RECDATE'] is not None:
                row['RECDATE'] = row['RECDATE'].strftime('%m-%d-%Y')
            if pm_number is not None:
                parts = pm_number.split(' ')
                row['CLIENT'] += ' (PM%s)' % parts[0]
                if len(parts) > 1:
                    row['CLIENT'] += ' %s' % ' '.join(parts[1:])
            if tract_number is not None:
                parts = tract_number.split(' ')
                row['CLIENT'] += ' (TR%s)' % parts[0]
                if len(parts) > 1:
                    row['CLIENT'] += ' %s' % ' '.join(parts[1:])
            row['NOTE'] = 'added #%d' % UPDATE_ID + (' - %s' % row['NOTE'] if row['NOTE'] else '')
            ws.append(row[key] for key in column_names)
        print('{0} rows added.'.format(cur.rowcount))

        print('Add MODIFIED map records to XLSX sheet "{0}" ...'.format(XLSX_SHEET_UPDATE_MODIFIED))

        ws = wb[XLSX_SHEET_UPDATE_MODIFIED]
        ws.delete_rows(2, ws.max_row - 1)
        column_names = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        cur.execute("""
            WITH q1 AS (
                -- aggregate an array of surveyor fullnames for each map in the update
                SELECT u.id map_id, array_agg(s.fullname) surveyors
                FROM (
                    SELECT id, regexp_split_to_table(upper(surveyors), '\s*&\s*') hollins_fullname
                    FROM {table_update_map}
                ) u
                -- assuming no missing surveyors here, LEFT JOIN ignores UNKNOWN, etc
                LEFT JOIN {table_update_surveyor} s ON s.hollins_fullname = u.hollins_fullname
                -- ignore new maps not already present in prod data
                JOIN {table_prod_map} m ON m.id = u.id
                GROUP BY u.id

            ), q2 AS (
                -- aggregate an array of surveyor fullnames for each map in prod
                SELECT m.id map_id, array_agg(s.fullname) surveyors
                FROM {table_prod_map} m
                LEFT JOIN {table_prod_signed_by} sb on sb.map_id = m.id
                LEFT JOIN {table_prod_surveyor} s on s.id = sb.surveyor_id
                GROUP BY m.id

            ), q3 AS (
                -- compare all fields except trs paths and note
                SELECT u.id map_id, u.maptype, u.book, u.page, u.lastpage - u.page + 1 npages,
                    u.recdate, array_to_string(ARRAY(SELECT unnest(q1.surveyors) ORDER BY 1), ', ') surveyors,
                    u.client, u.description
                FROM q1
                JOIN {table_update_map} u ON u.id = q1.map_id
                EXCEPT
                SELECT m.id, t.maptype, m.book, m.page, m.npages, recdate,
                    array_to_string(ARRAY(SELECT unnest(q2.surveyors) ORDER BY 1), ', ') surveyors,
                    regexp_replace(m.client, ' \((PM|TR)\d+\)( AMENDED| UNIT \d)?$', '') client,  description
                FROM q2
                JOIN {table_prod_map} m ON m.id = q2.map_id
                JOIN {table_prod_maptype} t ON m.maptype_id = t.id

            ), q4 AS (
                -- aggregate an array of trs paths for each map in the update
                SELECT u.map_id, array_agg(u.trs_path) paths
                FROM {table_update_trs_path} u
                GROUP BY map_id

                ), q5 AS (
                -- aggregate an array of trs paths for each map in prod
                SELECT map_id, array_agg(trs_path) paths
                FROM {table_prod_trs_path}
                GROUP BY map_id

            ), q6 AS (
                -- find maps with any trs path in the update that is not an ancestor of prod trs paths
                SELECT map_id, q4.paths paths
                FROM {table_update_trs_path} u
                JOIN q4 USING (map_id)
                JOIN q5 USING (map_id)
                WHERE NOT u.trs_path @> ANY (q5.paths)
                GROUP BY map_id, q4.paths

            )

            -- combine results of first comparison in q3 and trs paths comparison in q6
            SELECT u.id map_id, u.maptype, u.book, u.page, u.lastpage - u.page + 1 npages,
                coalesce(ARRAY(SELECT unnest(q4.paths) ORDER BY 1)::text[], ARRAY[]::text[]) trs_paths,
                u.recdate, u.surveyors, u.client, u.description, u.note,
                pm.pm_number, tr.tract_number
            FROM q3
            JOIN q4 ON q4.map_id = q3.map_id
            JOIN {table_update_map} u ON u.id = q3.map_id
            LEFT JOIN {table_update_pm_number} pm
                ON pm.maptype = u.maptype AND pm.book = u.book AND pm.page = u.page
            LEFT JOIN {table_update_tract_number} tr
                ON tr.maptype = u.maptype AND tr.book = u.book AND tr.page = u.page
            UNION
            SELECT u.id map_id, u.maptype, u.book, u.page, u.lastpage - u.page + 1 npages,
                coalesce(ARRAY(SELECT unnest(q6.paths) ORDER BY 1)::text[], ARRAY[]::text[]) trs_paths,
                u.recdate, u.surveyors, u.client, u.description, u.note,
                 pm.pm_number, tr.tract_number
            FROM q6
            JOIN {table_update_map} u ON u.id = q6.map_id
            LEFT JOIN {table_update_pm_number} pm
                ON pm.maptype = u.maptype AND pm.book = u.book AND pm.page = u.page
            LEFT JOIN {table_update_tract_number} tr
                ON tr.maptype = u.maptype AND tr.book = u.book AND tr.page = u.page
            ORDER BY map_id
            ;
        """.format(
            table_update_map=TABLE_UPDATE_MAP,
            table_update_trs_path=TABLE_UPDATE_TRS_PATH,
            table_update_surveyor=TABLE_UPDATE_SURVEYOR,
            table_update_pm_number=TABLE_UPDATE_PM_NUMBER,
            table_update_tract_number=TABLE_UPDATE_TRACT_NUMBER,
            table_prod_map=TABLE_PROD_MAP,
            table_prod_trs_path=TABLE_PROD_TRS_PATH,
            table_prod_surveyor=TABLE_PROD_SURVEYOR,
            table_prod_signed_by=TABLE_PROD_SIGNED_BY,
            table_prod_maptype=TABLE_PROD_MAPTYPE
        ))

        for row in cur:
            pm_number, tract_number = row[-2:]
            row = dict(zip(column_names, row))
            row['TRS_PATHS'] = '; '.join(abbrev_paths(row['TRS_PATHS']))
            if row['RECDATE'] is not None:
                row['RECDATE'] = row['RECDATE'].strftime('%m-%d-%Y')
            if pm_number is not None:
                parts = pm_number.split(' ')
                row['CLIENT'] += ' (PM%s)' % parts[0]
                if len(parts) > 1:
                    row['CLIENT'] += ' %s' % ' '.join(parts[1:])
            if tract_number is not None:
                parts = tract_number.split(' ')
                row['CLIENT'] += ' (TR%s)' % parts[0]
                if len(parts) > 1:
                    row['CLIENT'] += ' %s' % ' '.join(parts[1:])
            row['NOTE'] = 'modified #%d' % UPDATE_ID + (' - %s' % row['NOTE'] if row['NOTE'] else '')
            ws.append(row[key] for key in column_names)

            # ws.cell(ws.max_row, column_names.index('RECDATE') + 1).alignment =  Alignment(horizontal='left')

        print('{0} rows added.'.format(cur.rowcount))

        print('Add MISSING map records to XLSX sheet "{0}" ...'.format(XLSX_SHEET_UPDATE_MISSING))

        ws = wb[XLSX_SHEET_UPDATE_MISSING]
        ws.delete_rows(2, ws.max_row - 1)
        column_names = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        cur.execute("""
            WITH q1 AS (
                SELECT map_id, array_agg(trs_path) paths
                FROM {table_prod_trs_path}
                GROUP BY map_id
            )
            SELECT m.id map_id, t.maptype, m.book, m.page, m.npages,
                coalesce(q1.paths::text[], ARRAY[]::text[]) trs_paths,
                m.recdate, NULL, m.client, m.description, m.note
            FROM {table_prod_map} m
            JOIN {table_prod_maptype} t ON t.id = m.maptype_id
            LEFT JOIN q1 ON q1.map_id = m.id
            LEFT JOIN {table_update_map} u ON u.id = m.id
            WHERE u.id IS NULL
            ORDER BY m.id
            ;
        """.format(
            table_update_map=TABLE_UPDATE_MAP,
            table_prod_map=TABLE_PROD_MAP,
            table_prod_trs_path=TABLE_PROD_TRS_PATH,
            table_prod_maptype=TABLE_PROD_MAPTYPE
        ))

        for row in cur:
            row = dict(zip(column_names, row))
            row['TRS_PATHS'] = '; '.join(abbrev_paths(row['TRS_PATHS']))
            if row['RECDATE'] is not None:
                row['RECDATE'] = row['RECDATE'].strftime('%m-%d-%Y')
            row['NOTE'] = 'missing #%d' % UPDATE_ID + (' - %s' % row['NOTE'] if row['NOTE'] else '')
            ws.append(row[key] for key in column_names)

        print('{0} rows added.'.format(cur.rowcount))

        # Add changed trs paths to trs_paths sheet of XLSX update
        #
        # Due to parsed subsection (type 3) and other subsection paths added
        # to the TRS data outside the official Hollins update, the current TRS
        # path data cannot be directly compared to the update data. Instead we
        # ensure that for every TRS path in the update there is a descendant
        # or equal path in the current trs_path table. This ensures a full
        # section search returns the same results as an equivalent Hollins
        # search. Results from subsection searches however may differ.

        ws = wb[XLSX_SHEET_TRS_PATH]
        ws.delete_rows(2, ws.max_row - 1)

        print('Add MODIFIED TRS records to XLSX sheet "{0}" ...'.format(XLSX_SHEET_TRS_PATH))

        cur.execute("""
            WITH cur AS (
              SELECT map_id, array_agg(trs_path) paths
              FROM {table_prod_trs_path}
              GROUP BY map_id
            ), upd AS (
              SELECT map_id, array_agg(trs_path) paths
              FROM {table_update_trs_path}
              GROUP BY map_id
            )
            SELECT map_id,
              upd.paths::text[] upd,
              cur.paths::text[] cur,
              array_agg(trs_path)::text[] dif
            FROM {table_update_trs_path}
            JOIN cur USING (map_id)
            JOIN upd USING (map_id)
            WHERE NOT trs_path @> ANY (cur.paths)
            GROUP BY map_id, cur.paths, upd.paths
            ORDER BY map_id
            ;
        """.format(
            table_update_trs_path=TABLE_UPDATE_TRS_PATH,
            table_prod_trs_path=TABLE_PROD_TRS_PATH
        ))

        for map_id, *paths in cur:
            paths = ['; '.join(abbrev_paths(p)) for p in paths]
            ws.append((map_id, *paths))
        print('{0} rows added.'.format(cur.rowcount))

        wb.save(filename=XLSX_DATA_UPDATE)
        wb.close()

    print('Done.')


def cleanup_tables():

    print('\nCleanup tables ... ')

    print('DROP SCHEMA {}'.format(SCHEMA_UPDATE))
    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:
        cur.execute("""
            DROP SCHEMA {schema_update} CASCADE;
        """.format(schema_update=SCHEMA_UPDATE))

    print('Done.')

if __name__ == '__main__':

    startTime = time.time()

    load_update_tables()
    load_xlsx_tables()
    create_update()
    # cleanup_tables()

    endTime = time.time()
    print('\n{0:.3f} sec'.format(endTime - startTime))
