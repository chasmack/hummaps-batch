import os
import os.path
from glob import glob
from PIL import Image
from openpyxl import load_workbook
from subprocess import check_call
import psycopg2
import time
import re

from const import *

# test 2

def convert_maps():

    # Convert tiff map images in the scan directory to jpegs
    print('\nConverting map images...')

    # Turn of warnings about oversize image files
    Image.warnings.simplefilter('ignore', Image.DecompressionBombWarning)

    nfiles = 0

    for maptype in sorted(MAPTYPES.values()):

        for imagefile in sorted(glob(os.path.join(IMAGE_SCAN_DIR, maptype.lower(), '*',  '*.tif'))):

            src_path, src_file = os.path.split(imagefile.lower())
            dest_file = os.path.splitext(src_file)[0] + '.jpg'
            book = os.path.basename(src_path)

            dest_dir = os.path.join(IMAGE_MAP_DIR, maptype.lower(), book)
            os.makedirs(dest_dir, exist_ok=True)

            print('\n' + dest_file)

            with Image.open(imagefile) as img:

                print('Mode: %s' % ({
                    '1': '1-bit black and white',
                    'L': '8-bit greyscale',
                    'P': '8-bit color map',
                    'RGB': 'RGB color',
                    'RGBA': 'RGBa color'
                }[img.mode]))

                # Calculate the map image size
                scan_dpi = tuple(int(round(d)) for d in img.info['dpi'])
                map_dpi = tuple(min(MAP_DPI, dpi) for dpi in scan_dpi)
                map_size = tuple(int(round(d * mdpi / sdpi)) for d, mdpi, sdpi in zip(img.size, map_dpi, scan_dpi))

                print('Map size: %s @ %d dpi => %s @ %d dpi' % (str(img.size), scan_dpi[0], str(map_size), map_dpi[0]))

                # Convert 8-bit color map to RGB
                if img.mode == 'P':
                    print('Converting to RGB...')
                    img = img.convert('RGB')

                # Convert 1-bit black and white to 8-bit greyscale
                if img.mode == '1':
                    print('Converting to 8-bit greyscale...')
                    img = img.convert('L')

                # Resize and save the map image
                img = img.resize(map_size, resample=Image.BICUBIC)
                img.save(os.path.join(dest_dir, dest_file), dpi=map_dpi, quality=MAP_QUALITY)

                nfiles += 1

    print('\n%d files converted' % (nfiles))


def convert_ccs():

    # Convert tiff CC images in the scan directory to jpegs
    print('\nConverting cc images...\n')

    # Turn of warnings about oversize image files
    Image.warnings.simplefilter('ignore', Image.DecompressionBombWarning)

    dest_dir = os.path.join(IMAGE_MAP_DIR, 'cc')
    os.makedirs(dest_dir, exist_ok=True)

    nfiles = 0

    for imagefile in sorted(glob(os.path.join(IMAGE_SCAN_DIR, 'cc', '*.tif'))):

        src_path, src_file = os.path.split(imagefile.lower())

        # Only convert properly named CCs
        if not re.fullmatch('\d{4}-(or|doc)-\d{6}-\d{3}.tif', src_file):
            print('Bad CC filename: %s' % (src_file))
            continue

        dest_file = os.path.splitext(src_file)[0] + '.jpg'

        print('\n' + dest_file)

        with Image.open(imagefile) as img:

            print('Mode: %s' % ({
                '1': '1-bit black and white',
                'L': '8-bit greyscale',
                'P': '8-bit color map',
                'RGB': 'RGB color',
                'RGBA': 'RGBa color'
            }[img.mode]))

            # Calculate the map image size
            scan_dpi = tuple(int(round(d)) for d in img.info['dpi'])
            map_dpi = tuple(min(MAP_DPI, dpi) for dpi in scan_dpi)
            map_size = tuple(int(round(d * mdpi / sdpi)) for d, mdpi, sdpi in zip(img.size, map_dpi, scan_dpi))

            print('Map size: %s @ %d dpi => %s @ %d dpi' % (str(img.size), scan_dpi[0], str(map_size), map_dpi[0]))

            # Convert 8-bit color map to RGB
            if img.mode == 'P':
                print('Converting to RGB...')
                img = img.convert('RGB')

            # Convert 1-bit black and white to 8-bit greyscale
            if img.mode == '1':
                print('Converting to 8-bit greyscale...')
                img = img.convert('L')

            # Resize and save the map image
            img = img.resize(map_size, resample=Image.BICUBIC)
            img.save(os.path.join(dest_dir, dest_file), dpi=map_dpi, quality=MAP_QUALITY)

            nfiles += 1

    print('\n%d files converted' % (nfiles))


# Combine jpeg map images and any jpeg CC images into pdfs

def make_pdfs():

    print('\nCreating PDFs...\n')

    # Read the CC data from the XLSX workbook
    wb = load_workbook(filename=XLSX_DATA_UPDATE, read_only=True)
    ws = wb[XLSX_SHEET_CC]
    column_names = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    cc_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(column_names, row))

        if row['MAPTYPE'] in MAPTYPES:
            map = '%03d%s%03d' % (row['BOOK'], MAPTYPES[row['MAPTYPE']].lower(), row['PAGE'])
        else:
            print('ERROR: Bad MAPTYPE: %s' % row['MAPTYPE'])
            exit(1)

        m1 = re.fullmatch('(\d+) OR (\d+)', row['DOC_NUMBER'])
        m2 = re.fullmatch('(\d{4})-(\d+)-\d+', row['DOC_NUMBER'])
        if (m1):
            row['DOC_NUMBER'] = '{0:04d}-or-{1:06d}'.format(*(int(n) for n in m1.groups()))
        elif (m2):
            row['DOC_NUMBER'] = '{0:04d}-doc-{1:06d}'.format(*(int(n) for n in m2.groups()))
        else:
            print('ERROR: CC doc number formant error: "%s"' % row['DOC_NUMBER'])
            exit(1)

        cc_data[map] = cc_data.get(map, []) + [row]

    # List of maps with missing CCs
    missing_ccs = []

    nfiles = 0
    for maptype in sorted(MAPTYPES.values()):

        # Get a list of map images indexed by map name
        mapimages = {}
        for imagefile in sorted(glob(os.path.join(IMAGE_MAP_DIR, maptype, '*', '*.jpg'))):
            map = re.match('^(\d{3}\D{2}\d{3})', os.path.basename(imagefile)).groups()[0]
            mapimages[map] = mapimages.get(map, []) + [imagefile]


        for map, imagefiles in sorted(mapimages.items(), key=lambda k: k[0]):

            # Append CC images to the end of the map image list
            if maptype.upper() in ('PM', 'RM', 'RS') and map in cc_data:
                for cc in sorted(cc_data[map], key=lambda cc: cc['DOC_NUMBER'], reverse=True):
                    ccimages = sorted(glob(os.path.join(IMAGE_MAP_DIR, 'cc', cc['DOC_NUMBER'] + '-*.jpg')))
                    if len(ccimages) == 0:
                        print('WARNING: Missing CC for %s: %s' % (map, cc['DOC_NUMBER']))
                        missing_ccs.append(cc['DOC_NUMBER'])
                    elif len(ccimages) != cc['NPAGES']:
                        print('ERROR: Bad CC page count for %s: %s' % (map, cc['DOC_NUMBER']))
                        exit(1)
                    else:
                        mapimages[map] += ccimages

            book = map[0:3]
            dest = os.path.join(IMAGE_PDF_DIR, maptype.lower(), book, map + '.pdf')
            os.makedirs(os.path.dirname(dest), exist_ok=True)

            magick_cmd = [MAGICK] + imagefiles + [dest]

            print('%s: %s' % (map, dest))
            check_call(magick_cmd)
            nfiles += 1

    print('\n%d PDFs created.' % nfiles)
    if (missing_ccs):
        print('Missing CCs: %s' % ', '.join(missing_ccs))


def update_map_image():

    # Update the map image table to link map images to map records

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        print('INSERT INTO {table_prod_map_image} ...'.format(table_prod_map_image=TABLE_PROD_MAP_IMAGE))

        # Invert the mapping from maptype to maptype abbrev.
        maptype_name = dict((v.lower(), k) for k, v in MAPTYPES.items())

        # Create a list of imagefiles indexed by map name.
        map_imagefiles = {}
        for abbrev in sorted(maptype_name.keys()):
            for book in (path[-3:] for path in sorted(glob(os.path.join(IMAGE_MAP_DIR, abbrev, '*')))):
                for imagefile in sorted(glob(os.path.join(IMAGE_MAP_DIR, abbrev, book, '*'))):
                    imagefile = '/map/%s/%s/%s' % (abbrev, book, os.path.basename(imagefile))
                    map_name = re.match('.*/(.*)-', imagefile).group(1)
                    map_imagefiles[map_name] = map_imagefiles.get(map_name, []) + [imagefile]

        # Rearrange into a list of dictionary objects for the query.
        imagefiles = []
        for map_name in sorted(map_imagefiles.keys()):
            book, abbrev, page = re.match('(\d+)(\D+)(\d+)', map_name).groups()
            imagefiles.append(dict((
                ('ABBREV', abbrev.upper()),
                ('BOOK', int(book)),
                ('PAGE', int(page)),
                ('IMAGEFILES', map_imagefiles[map_name])
            )))

        cur.executemany("""
            WITH q1 AS (
              SELECT m.id map_id
              FROM {table_prod_map} m
              JOIN {table_prod_maptype} t ON t.id = m.maptype_id
              WHERE t.abbrev = %(ABBREV)s AND m.book = %(BOOK)s AND m.page = %(PAGE)s
            ), d1 AS (
              DELETE FROM {table_prod_map_image}
              WHERE map_id IN (SELECT map_id FROM q1)
            )
            INSERT INTO {table_prod_map_image} (map_id, imagefile, page)
            SELECT map_id, imagefile,
              regexp_replace(imagefile, '.*-(\d+).*', '\\1')::int page
            FROM q1, (
              SELECT unnest(%(IMAGEFILES)s) imagefile
            ) q2
            ;
        """.format(
            table_prod_map_image=TABLE_PROD_MAP_IMAGE,
            table_prod_map=TABLE_PROD_MAP,
            table_prod_maptype=TABLE_PROD_MAPTYPE
        ), imagefiles)
        con.commit()
        print('%d rows effected.' % cur.rowcount)


def update_cc_image():

    # Update the cc_image table to link CC images to records in the cc table
    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        print('INSERT INTO {table_prod_cc_image} ...'.format(table_prod_cc_image=TABLE_PROD_CC_IMAGE))

        # Create a list of imagefiles indexed by CC name.
        cc_imagefiles = {}
        for imagefile in sorted(glob(os.path.join(IMAGE_MAP_DIR, 'cc', '*'))):
            imagefile = '/map/cc/%s' % os.path.basename(imagefile)
            cc_name = re.match('.*/(\d+-\D+-\d+)', imagefile).group(1)
            cc_imagefiles[cc_name] = cc_imagefiles.get(cc_name, []) + [imagefile]

        # Rearrange into a list of dictionary objects for the query.
        imagefiles = []
        for cc_name in sorted(cc_imagefiles.keys()):
            book, page = re.match('.*/0*(\d+)-\D+-0*(\d+)', imagefile).groups()
            imagefiles.append(dict((
                ('BOOK', book),
                ('PAGE', page),
                ('IMAGEFILES', cc_imagefiles[cc_name])
            )))

        cur.executemany("""
            WITH q1 AS (
              SELECT cc.id cc_id
              FROM {table_prod_cc} cc
              WHERE cc.doc_number SIMILAR TO %(BOOK)s || '(-0*| OR )' || %(PAGE)s || '%%'
            ), d1 AS (
              DELETE FROM {table_prod_cc_image}
              WHERE cc_id IN (SELECT cc_id FROM q1)
            )
            INSERT INTO {table_prod_cc_image} (cc_id, imagefile, page)
            SELECT cc_id, imagefile,
              regexp_replace(imagefile, '.*-(\d+).*', '\\1')::int page
            FROM q1, (
              SELECT unnest(%(IMAGEFILES)s) imagefile
            ) q2
            ;
        """.format(
            table_prod_cc=TABLE_PROD_CC,
            table_prod_cc_image=TABLE_PROD_CC_IMAGE
        ), imagefiles)

        print('%d rows effected.' % cur.rowcount)


def update_pdf():

    # Update the pdf table to link pdf files to map records
    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        print('INSERT INTO {table_prod_pdf} ...'.format(table_prod_pdf=TABLE_PROD_PDF))

        # Invert the mapping from maptype to maptype abbrev.
        maptype_name = dict((v.lower(), k) for k, v in MAPTYPES.items())

        # Create a list of dictionary objects for the query.
        pdffiles = []
        for abbrev in sorted(maptype_name.keys()):
            for book in (path[-3:] for path in sorted(glob(os.path.join(IMAGE_PDF_DIR, abbrev, '*')))):
                for pdffile in sorted(glob(os.path.join(IMAGE_PDF_DIR, abbrev, book, '*'))):
                    pdffile = '/pdf/%s/%s/%s' % (abbrev, book, os.path.basename(pdffile))
                    page = re.match('.*/.*\D(\d+)', pdffile).group(1)
                    pdffiles.append(dict((
                        ('ABBREV', abbrev.upper()),
                        ('BOOK', int(book)),
                        ('PAGE', int(page)),
                        ('PDFFILE', pdffile)
                    )))

        cur.executemany("""
            WITH q1 AS (
              SELECT m.id map_id
              FROM {table_prod_map} m
              JOIN {table_prod_maptype} t ON t.id = m.maptype_id
              WHERE t.abbrev = %(ABBREV)s AND m.book = %(BOOK)s AND m.page = %(PAGE)s
            ), d1 AS (
              DELETE FROM {table_prod_pdf}
              WHERE map_id IN (SELECT map_id FROM q1)
            )
            INSERT INTO {table_prod_pdf} (map_id, pdffile)
            SELECT map_id, %(PDFFILE)s
            FROM q1
            ;
        """.format(
            table_prod_pdf=TABLE_PROD_PDF,
            table_prod_map=TABLE_PROD_MAP,
            table_prod_maptype=TABLE_PROD_MAPTYPE
        ), pdffiles)
        con.commit()
        print('%d rows effected.' % cur.rowcount)


if __name__ == '__main__':

    startTime = time.time()

    convert_maps()
    convert_ccs()
    make_pdfs()
    update_map_image()
    update_cc_image()
    update_pdf()

    endTime = time.time()
    print('\n{0:.3f} sec'.format(endTime - startTime))


