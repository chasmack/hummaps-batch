from openpyxl import load_workbook
from dateutil.parser import parse
import datetime
from trs_path import expand_paths, abbrev_paths
import re

from const import *


def hollins_trs(save_wb=False):

    # Map hollins subsection numbers to trs_path subsection codes
    subsec_code = dict(zip(range(1, 37), 'DCBAEFGHLKJIMNOP'))

    wb = load_workbook(XLSX_DATA_MAP)

    # Get surveyor fullnames indexed by hollins fullname
    hollins_fullname_to_fullname = {}
    ws = wb['surveyor']
    xl_cols = list(c.value for c in ws[1])
    for xl_row in ws.iter_rows(min_row=2):
        hollins_rec = dict(zip((k.lower() for k in xl_cols), (c.value for c in xl_row)))
        hollins_fullname_to_fullname[hollins_rec['hollins_fullname']] = hollins_rec['fullname']

    # Extract trs path specs form Hollins TRS indexed by map id
    hollins_paths = {}

    ws = wb['hollins_trs']
    xl_cols = list(c.value for c in ws[1])
    for xl_row in ws.iter_rows(min_row=2):
        hollins_rec = dict(zip((k.lower() for k in xl_cols), (c.value for c in xl_row)))
        map_id = hollins_rec['id']
        for sec in hollins_rec['section'].strip(',').split(','):
            path = '.'.join((hollins_rec['township'].lstrip('0'), hollins_rec['range'], sec))
            hollins_paths.setdefault(map_id, []).append(path)

    # Copy Hollins map data into the update map records
    update_recs = []

    ws = wb['hollins_map']
    xl_cols = list(c.value for c in ws[1])
    for xl_row in ws.iter_rows(min_row=2):
        hollins_rec = dict(zip((k.lower() for k in xl_cols), (c.value for c in xl_row)))
        update_rec = {}

        # Create an update record from hollins data
        update_rec['map_id'] = hollins_rec['id']
        update_rec['maptype'] = hollins_rec['maptype']
        update_rec['book'] = hollins_rec['book']
        update_rec['page'] = hollins_rec['firstpage']
        update_rec['npages'] = hollins_rec['lastpage'] - hollins_rec['firstpage'] + 1

        if type(hollins_rec['recdate']) is datetime.datetime:
            dt = hollins_rec['recdate']
            update_rec['recdate'] = '%d/%d/%d' % (dt.month, dt.day, dt.year)
        else:
            update_rec['recdate'] = hollins_rec['recdate']


        if hollins_rec['surveyor'] == 'UNKNOWN':
            update_rec['surveyors'] = None

        else:
            surveyors = []
            for hollins_fullname in re.split('\s*&\s*', hollins_rec['surveyor']):
                if hollins_fullname in hollins_fullname_to_fullname:
                    surveyors.append(hollins_fullname_to_fullname[hollins_rec['surveyor']])
                else:
                    print('WARNING [%d]: Hollins surveyor not found: %s' % (xl_row[0].row, hollins_rec['surveyor']))

            update_rec['surveyors'] = ', '.join(surveyors) if surveyors else None

        update_rec['client'] = hollins_rec['donefor']
        update_rec['description'] = hollins_rec['descrip']
        update_rec['note'] = hollins_rec['comment']

        # Search the column name list for subsection columns
        map_id = hollins_rec['id']
        trs_paths = hollins_paths.setdefault(map_id, [])
        for col in xl_cols:
            m = re.fullmatch('0?(\d(?:N|S))(\d(?:E|W))0?(\d+)', col)
            col = col.lower()
            if not (m and hollins_rec[col]):
                continue
            trs = '.'.join(m.groups())

            # Remove the ancestor trs path record
            if trs in trs_paths:
                trs_paths.remove(trs)
            else:
                print('WARNING: No Hollins trs record for subsection: map_id=%d: trs=%s' % (map_id, trs))

            # Add the subsection paths
            hollins_subsecs = map(int, hollins_rec[col].strip(',').split(','))
            trs_paths += ['%s.%s' % (trs, subsec_code[subsec]) for subsec in hollins_subsecs]

        update_rec['trs_paths'] = '; '.join(abbrev_paths(trs_paths)) if trs_paths else None

        update_recs.append(update_rec)

    # Add map records to the end of update
    ws = wb['update']
    xl_cols = list(c.value for c in ws[1])
    for update_rec in update_recs:
        ws.append(update_rec[col.lower()] for col in xl_cols)

    if save_wb:
        wb.save(filename=XLSX_DATA_MAP)
    wb.close()


if __name__ == '__main__':

    hollins_trs(save_wb=False)
