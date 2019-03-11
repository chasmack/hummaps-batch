import psycopg2
import requests
from openpyxl import load_workbook
from dateutil.parser import parse
from trs_path import expand_paths, abbrev_paths, trs_path_sortkey
import re

from const import *

def check_surveyor():

    with psycopg2.connect(DSN_PROD) as con, con.cursor() as cur:

        # Get the surveyors and hollins aliases
        surveyors = []
        wb = load_workbook(XLSX_DATA_SURVEYOR)
        xl = wb.active
        ws_cols = list(c.value for c in xl[1])
        for ws_row in xl.iter_rows(min_row=2):
            xl = dict(zip((v.lower() for v in ws_cols), (v.value for v in ws_row)))
            xl_cell = dict(zip((v.lower() for v in ws_cols), ws_row))
            fullname = [xl['firstname'], xl['lastname']]
            if xl['secondname']:
                fullname.insert(-1, xl['secondname'][0])
            if xl['thirdname']:
                fullname.insert(-1, xl['thirdname'][0])
            if xl['suffix']:
                fullname.append(xl['suffix'])
            xl_cell['fullname'].value = ' '.join(fullname)
        wb.close()

    # wb.save(filename=XLSX_DATA_SURVEYOR)
    wb.close()



if __name__ == '__main__':

    check_surveyor()
