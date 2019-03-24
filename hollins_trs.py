from openpyxl import load_workbook
from dateutil.parser import parse
import datetime
from trs_path import expand_paths, abbrev_paths
import re

from const import *


def hollins_trs():

    # Map hollins subsection numbers to trs_path subsection codes
    HOLLINS_SUBSEC = dict(zip(range(1, 37), 'DCBAEFGHLKJIMNOP'))

    do_save = False
    wb = load_workbook(XLSX_DATA_MAP)

    hollins_paths = {}
    for ws_row in wb['trs']:
        map_id, tshp, rng, secs = list(c.value for c in ws_row)
        if map_id not in hollins_paths:
            hollins_paths[map_id] = {}
        for sec in secs.strip(',').split(','):
            trs = '.'.join([tshp.lstrip('0'), rng, sec])
            hollins_paths[map_id][trs] = []

    ws = wb['subsec']
    ws_cols = list(c.value for c in ws[1])
    ws_cols[0] = 'map_id'
    for i in range(1, len(ws_cols)):
        ws_cols[i] = '.'.join(re.fullmatch('0?(\d(?:N|S))(\d(?:E|W))0?(\d+)', ws_cols[i]).groups())

    for ws_row in ws.iter_rows(min_row=2):
        map_id = ws_row[0].value
        if map_id not in hollins_paths:
            hollins_paths[map_id] = {}
        for trs, subsecs in zip(ws_cols[1:], (v.value for v in ws_row[1:])):
            if not subsecs:
                continue
            if trs not in hollins_paths[map_id]:
                hollins_paths[map_id][trs] = []
            for ss in subsecs.strip(',').split(','):
                hollins_paths[map_id][trs].append(HOLLINS_SUBSEC[int(ss)])

    for map_id in sorted(hollins_paths.keys()):
        paths = []
        for trs in hollins_paths[map_id].keys():
            if len(hollins_paths[map_id][trs]) == 0:
                paths.append(trs)
            else:
                paths += ['%s.%s' % (trs, ss) for ss in hollins_paths[map_id][trs]]
        hollins_paths[map_id] = abbrev_paths(paths)

    ws = wb['update']
    ws_cols = list(c.value for c in ws[1])

    for ws_row in ws.iter_rows(min_row=2):
        if all(c.value is None for c in ws_row):
            continue

        ws_cell = dict(zip((v.lower() for v in ws_cols), ws_row))
        map_id = ws_cell['map_id'].value
        update_paths = ws_cell['trs_paths'].value

        if map_id in hollins_paths and update_paths:
            update_paths = re.split(';?\s+', update_paths)
            if expand_paths(update_paths) != expand_paths(hollins_paths[map_id]):
                update = '; '.join(update_paths)
                hollins = '; '.join(hollins_paths[map_id])
                print('map_id=%d: update: %s  hollins: %s' % (map_id, update, hollins))
        elif map_id in hollins_paths:
            ws_cell['trs_paths'].value = '; '.join(hollins_paths[map_id])
            do_save = True

    if do_save:
        wb.save(filename=XLSX_DATA_MAP)
    wb.close()


if __name__ == '__main__':

    hollins_trs()
